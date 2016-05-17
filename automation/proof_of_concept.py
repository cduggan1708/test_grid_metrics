import sys, getopt
from datetime import datetime
import openpyxl
from MemberMetricData2 import MemberMetricData2
from MetricSelectorMapper import MetricSelectorMapper

column_letter = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
                 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S',
                 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z', 27: 'AA', 28: 'AB', 
                 29: 'AC', 30: 'AD', 31: 'AE', 32: 'AF', 33: 'AG', 34: 'AH', 35: 'AI', 36: 'AJ', 37: 'AK', 
                 38: 'AL', 39: 'AM', 40: 'AN', 41: 'AO', 42: 'AP', 43: 'AQ', 44: 'AR', 45: 'AS', #etc
                 # specific values that are hardcoded (based on template)
                 'MemberID': 'B', # test grid
                 'ProgramName': 'A', 'MetricID': 'B', 'MetricDisplayName': 'C', 'MetricSelector': 'D' # mapper
                 }

def parseTestGridIntoObjects(test_grid_file):
    return readMetricDataFromTestGrid(test_grid_file)

def parseMetricSelectorMapperIntoObjects(metric_selector_mapper):
    return readMapperDataFromMetricSelectorMapper(metric_selector_mapper)

def readMetricDataFromTestGrid(test_grid_file):  
    first_row, last_row, last_column, first_metric_column, workbook = getRelevantCellsInTestGrid(test_grid_file)
    grid_sheet = workbook.get_sheet_by_name('Grid')

    member_metric_data_list = []

    # iterate rows
    for row in range(first_row + 2, last_row + 1):
        member_id = grid_sheet[column_letter['MemberID'] + str(row)].value
        # if member id is None, there is nothing to insert
        if member_id is not None:

            # iterate columns
            for column in range(first_metric_column, last_column + 1):        
                #print("Cell %s%d\n" % (column_letter[column], row))
                metric_value = grid_sheet[column_letter[column] + str(row)].value

                # if metric value is None, there is nothing to insert
                if metric_value is not None:
                    metric_id = grid_sheet[column_letter[column] + str(first_row)].value

                    # if metric id is None, there is nothing to insert
                    if metric_id is not None:

                        # rule: if value is a string, set metric data type to refcode
                        if isinstance(metric_value, str):
                            metric_data_type = 'refcode'
                        else:
                            metric_data_type = grid_sheet[column_letter[column] + str(first_row + 1)].value

                        # if metric data type is None, there is nothing to insert
                        if metric_data_type is not None:

                            # create the object and append to list
                            mmd = MemberMetricData2(member_id, metric_id, metric_data_type, metric_value)
                            member_metric_data_list.append(mmd)

    return member_metric_data_list

def readMapperDataFromMetricSelectorMapper(metric_selector_mapper):
    first_row, last_row, last_column, first_column, workbook = getRelevantCellsInMetricSelectorMapper(metric_selector_mapper)
    mapper_sheet = workbook.get_sheet_by_name('Mapper')

    metric_selector_data_list = []

    # iterate rows
    for row in range(first_row + 1, last_row + 1):
        program_name = mapper_sheet[column_letter['ProgramName'] + str(row)].value
        if program_name is not None:
            metric_id = mapper_sheet[column_letter['MetricID'] + str(row)].value
            if metric_id is not None:
                metric_display_name = mapper_sheet[column_letter['MetricDisplayName'] + str(row)].value
                if metric_display_name is not None:
                    metric_top_level_selector = mapper_sheet[column_letter['MetricSelector'] + str(row)].value
                    if metric_top_level_selector is not None:

                        # handle grouped metrics
                        metrics = str(metric_id).split('/')
                        for metric in metrics:
                        
                            # create the object and append to list
                            msm = MetricSelectorMapper(metric, metric_display_name, program_name, metric_top_level_selector)
                            metric_selector_data_list.append(msm)

    return metric_selector_data_list

def getRelevantCellsInTestGrid(test_grid_file):

    workbook = openpyxl.load_workbook(test_grid_file)
    grid_sheet = workbook.get_sheet_by_name('Grid')
    last_row = grid_sheet.max_row
    first_row = 1
    i = 0
    for row in range(1, last_row):
        value = grid_sheet['A' + str(row)].value
        i += 1
        if value is not None and 'MetricID' in value: # this is the first cell we care about
            first_row = i
            break

    last_column = grid_sheet.max_column
    first_metric_column = 0
    j = 0

    for column in range(1, last_column):
        value = grid_sheet[column_letter[column] + str(first_row)].value
        j += 1

        if first_metric_column == 0 and isinstance(value, int):
            first_metric_column = j # first int we found, meaning first metric id

        if first_metric_column != 0 and value is None: # should this check that not type int instead or in addition to?
            break

    #update last column to the one we care about
    last_column = j - 1

    print("Test Grid cells: A%d through %s\n" % (first_row, column_letter[last_column] + str(last_row)))

    return (first_row, last_row, last_column, first_metric_column, workbook)

def getRelevantCellsInMetricSelectorMapper(metric_selector_mapper):
    workbook = openpyxl.load_workbook(metric_selector_mapper)
    mapper_sheet = workbook.get_sheet_by_name('Mapper')
    last_row = mapper_sheet.max_row
    first_row = 1
    last_column = column_letter[4] # D
    first_column = 1

    print("Metric Selector Mapper cells: A%d through %s\n" % (first_row, column_letter[4] + str(last_row)))

    return (first_row, last_row, last_column, first_column, workbook)

def main(argv):
    # mmd = MemberMetricData2(100, 1000, 'float', '25')
    # mmd = MemberMetricData2(100, 1001, 'float', '150')
    # mmd = MemberMetricData2(101, 1000, 'float', '25')

    # print(mmd.find_by_member_metric_ids(100, 1000))

    # items = mmd.find_by_member_metric_ids(100, 1000)
    # for item in items:
    #     item.print()

    # metric_id = '10151/10140'
    # metrics = metric_id.split('/')
    # for metric in metrics:
    #     print(metric)

    # msm = MetricSelectorMapper(1000, 'Metric Name', 'Program Name', '#im_campaign_1084 > div.incentiveActivities > div:nth-child(2) > div > ul:nth-child(2)')
    # msm.print()

    test_grid_file = ''
    metric_selector_mapper = ''
    try:
        opts, args = getopt.getopt(argv, "ht:m:")
    except:
        print('test_grid_metrics.py -t <test grid filename> -m <metric selector mapper>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('test_grid_metrics.py -t <test grid filename> -m <metric selector mapper>')
            sys.exit()
        elif opt in ('-t', '--testGridFile'):
            test_grid_file = arg

        elif opt in ('-m', '--metricSelectorMapper'):
            metric_selector_mapper = arg

    print("%s: Executed proof_of_concept.py\n" % datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    if test_grid_file is not '':
        test_grid_data = parseTestGridIntoObjects(test_grid_file)
        for item in test_grid_data:
            item.print()
            print('\n')

        if metric_selector_mapper is not '':
            metric_selector_data = parseMetricSelectorMapperIntoObjects(metric_selector_mapper)
            for item in metric_selector_data:
                item.print()
                # test code
                # mmd = MemberMetricData2
                # items = mmd.find_by_member_metric_ids(7250433, 10151)
                # for item in items:
                #     item.print()
                # sys.exit(2)
                print('\n')

if __name__ == '__main__':
    main(sys.argv[1:])