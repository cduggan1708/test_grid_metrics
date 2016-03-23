import sys, getopt
from datetime import datetime
import os
import openpyxl
from MemberMetricData import MemberMetricData

column_letter = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J',
                 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S',
                 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z', 27: 'AA', 28: 'AB', 
                 29: 'AC', 30: 'AD', 31: 'AE', 32: 'AF', 33: 'AG', 34: 'AH', 35: 'AI', 36: 'AJ', 37: 'AK', 
                 38: 'AL', 39: 'AM', 40: 'AN', 41: 'AO', 42: 'AP', 43: 'AQ', 44: 'AR', 45: 'AS', #etc
                 # specific values that are hardcoded (based on template)
                 'MemberID': 'B'}

metric_type_query = {'float': "exec [LifeChanging].[METD].[usp_InsertFloatValueActivities] @memberID = %d, @MetricID = %d, @Float = %s, @measured=@today,@CreatedAt=@today",
                     'boolean': "exec [LifeChanging].[METD].[usp_InsertBoolValueActivities] @memberID = %d, @MetricID = %d, @Bool = %s, @measured=@today,@CreatedAt=@today",
                     'enum': "exec [LifeChanging].[METD].[usp_InsertEnumValueActivities] @memberID = %d, @MetricID = %d, @EnumValueID = %s, @measured=@today,@CreatedAt=@today",
                     # use yesterday for reference codes so that if the member also have float/boolean/enum data for today, a separate record will be created for the reference code
                     'refcode': "exec [LifeChanging].[METD].[usp_InsertReferenceCodeActivities] @memberID = %d, @MetricID = %d, @ReferenceCodeID = %s, @measured=@yesterday,@CreatedAt=@yesterday"}

reference_code_id = {'DNP': 1, 'N/A': 2, 'P': 3, 'UTO': 4, 'QNS': 5, 'NSA': 6, 'RDNO': 7, 'RDNOX2': 8, 'NULL': 9, 'E4': 10, 'DSO': 11}

def writeMemberMetricDataWithInsertsToNewSheet(member_metric_data_list, workbook, test_grid_file):
    sheet = workbook.create_sheet(title='InsertData')

    i = 1
    for mmd in member_metric_data_list:
        member_id = mmd.getMemberId()
        metric_id = mmd.getMetricId()
        metric_value = mmd.getMetricValue()
        metric_data_type = mmd.getMetricDataType()
        if metric_data_type == 'refcode':
            metric_value = reference_code_id[metric_value]

        sheet['A' + str(i)] = member_id
        sheet['B' + str(i)] = metric_id
        sheet['C' + str(i)] = metric_value
        sheet['D' + str(i)] = metric_type_query[metric_data_type] % (member_id, metric_id, str(metric_value))

        i += 1

    workbook.save(test_grid_file)

def readMetricDataFromTestGrid(test_grid_file):  
    first_row, last_row, last_column, first_metric_column, workbook = getRelevantCellsInTestGrid(test_grid_file)
    grid_sheet = workbook.get_sheet_by_name('Grid')

    member_metric_data_list = []

    # iterate rows
    for row in range(first_row + 2, last_row + 1):
        member_id = grid_sheet[column_letter['MemberID'] + str(row)].value
        # if member id is None, there is nothing to insert
        if member_id is not None:

            # iterate columns
            for column in range(first_metric_column, last_column + 1):        
                #print("Cell %s%d\n" % (column_letter[column], row))
                metric_value = grid_sheet[column_letter[column] + str(row)].value

                # if metric value is None, there is nothing to insert
                if metric_value is not None:
                    metric_id = grid_sheet[column_letter[column] + str(first_row)].value

                    # if metric id is None, there is nothing to insert
                    if metric_id is not None:

                        # rule: if value is a string, set metric data type to refcode
                        if isinstance(metric_value, str):
                            metric_data_type = 'refcode'
                        else:
                            metric_data_type = grid_sheet[column_letter[column] + str(first_row + 1)].value

                        # if metric data type is None, there is nothing to insert
                        if metric_data_type is not None:
                            mmd = MemberMetricData()
                            mmd.setMemberId(member_id)
                            mmd.setMetricId(metric_id)
                            mmd.setMetricDataType(metric_data_type)
                            mmd.setMetricValue(metric_value)
                            member_metric_data_list.append(mmd)

    writeMemberMetricDataWithInsertsToNewSheet(member_metric_data_list, workbook, test_grid_file)


def getRelevantCellsInTestGrid(test_grid_file):

    workbook = openpyxl.load_workbook(test_grid_file)
    grid_sheet = workbook.get_sheet_by_name('Grid')
    last_row = grid_sheet.max_row
    first_row = 1
    i = 0
    for row in range(1, last_row):
        value = grid_sheet['A' + str(row)].value
        i += 1
        if value is not None and 'MetricID' in value: # this is the first cell we care about
            first_row = i
            break

    last_column = grid_sheet.max_column
    first_metric_column = 0
    j = 0

    for column in range(1, last_column):
        value = grid_sheet[column_letter[column] + str(first_row)].value
        j += 1

        if first_metric_column == 0 and isinstance(value, int):
            first_metric_column = j # first int we found, meaning first metric id

        if first_metric_column != 0 and value is None: # should this check that not type int instead or in addition to?
            break

    #update last column to the one we care about
    last_column = j - 1

    print("Test Grid cells: A%d through %s" % (first_row, column_letter[last_column] + str(last_row)))

    return (first_row, last_row, last_column, first_metric_column, workbook)

def main(argv):
    test_grid_file = ''
    try:
        opts, args = getopt.getopt(argv, "hf:")
    except:
        print('test_grid_metrics.py -f <test grid filename>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('test_grid_metrics.py -f <test grid filename>')
            sys.exit()
        else:
            if opt in ('-f', '--testGridFile'):
                test_grid_file = arg
    
    # verify extension is excel file
    filename, extension = os.path.splitext(test_grid_file)
    if test_grid_file == '' or "xls" not in extension:
        print('Usage: test_grid_metrics.py -f <test grid filename>')
        sys.exit()

    print("%s: Executed test_grid_metrics.py" % datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    readMetricDataFromTestGrid(test_grid_file)

if __name__ == '__main__':
    main(sys.argv[1:])